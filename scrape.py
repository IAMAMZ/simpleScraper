import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

book_titles= []
book_prices=[]


URL = "http://books.toscrape.com/catalogue/category/books_1/index.html"



def scrapePrice(tag :object)->bool:
    if tag.has_attr("class") and tag['class'] == "price_color":
        return True
    return False


r = requests.get(URL)
soup = BeautifulSoup(r.content,"html5lib");
h3 = soup.find_all('h3')

for item in h3:
    for child in item:
        book_titles.append(child.string)


content = soup.body.find_all(attrs={"class":"price_color"})

for item in content:
    for child in item:
        book_prices.append(child.string)

#now combine the two lists

book_price_title = []

for i in range(len(book_titles)):
    book_price_title.append([book_titles[i],book_prices[i]])

#make a workbook
wb = Workbook()

#get the active workbook
ws = wb.active
ws["A1"] = "Book Title";
ws["B1"] = "Book Price"

for i in range(len(book_price_title)):
    print(book_price_title[i])
    bookTitle = book_price_title[i][0]
    bookPrice = book_price_title[i][1]
    ws[f"A{i+2}"] = bookTitle
    ws[f"B{i+2}"] = bookPrice


wb.save("book_prices.xlsx")